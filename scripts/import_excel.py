#!/usr/bin/env python3
"""
import_excel.py — Bibliothèque de prompts
Lit prompt_lib_import_template.xlsx (ou CSV) → écrit import_queue.json

Usage:
  python scripts/import_excel.py <fichier.xlsx|csv>
  python scripts/import_excel.py <fichier.xlsx|csv> --replace
"""

import sys
import json
import uuid
import re
import argparse
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl requis : pip install openpyxl")

# ── Config ──────────────────────────────────────────────────────────────
ROOT         = Path(__file__).resolve().parent.parent
OUTPUT_PATH  = ROOT / "import_queue.json"
VALID_TYPES  = {"prompt_texte", "prompt_image", "tuto", "skill"}
SKIP_ROWS    = 2          # ligne 1 = headers, ligne 2 = notes → données à partir de 3

# ── Helpers ─────────────────────────────────────────────────────────────
def make_id():
    return str(uuid.uuid4())

def now_iso():
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

def parse_tags(raw):
    if not raw:
        return []
    return [t.strip() for t in re.split(r"[,;]", str(raw)) if t.strip()]

def clean(val):
    if val is None:
        return ""
    return str(val).strip()

def is_empty_row(row_values):
    return all(v is None or str(v).strip() == "" for v in row_values)

# ── Parsers ─────────────────────────────────────────────────────────────
def parse_entries_sheet(ws):
    entries = []
    errors  = []
    cols = ["id","titre","contenu","type","tags","notes","thumb_path","parent_id"]

    for row_idx, row in enumerate(ws.iter_rows(min_row=SKIP_ROWS+1, values_only=True), start=SKIP_ROWS+1):
        if is_empty_row(row):
            continue

        r = dict(zip(cols, [clean(v) for v in row[:len(cols)]]))

        # Validation type
        if r["type"] and r["type"] not in VALID_TYPES:
            errors.append(f"  [entries] ligne {row_idx} : type invalide '{r['type']}' — attendu : {', '.join(sorted(VALID_TYPES))}")
            continue

        # Titre obligatoire
        if not r["titre"]:
            errors.append(f"  [entries] ligne {row_idx} : 'titre' vide — ligne ignorée")
            continue

        entry = {
            "id":            r["id"] or make_id(),
            "titre":         r["titre"],
            "contenu":       r["contenu"],
            "type":          r["type"] or "prompt_texte",
            "tags":          parse_tags(r["tags"]),
            "notes":         r["notes"],
            "thumb_path":    r["thumb_path"],
            "date_creation": now_iso(),
            "parent_id":     r["parent_id"] or None,
        }
        entries.append(entry)

    return entries, errors


def parse_briques_sheet(ws):
    briques = []
    errors  = []
    cols = ["id","categorie","valeur","tags"]
    valid_cats = {"style","sujet","qualité","ambiance"}

    for row_idx, row in enumerate(ws.iter_rows(min_row=SKIP_ROWS+1, values_only=True), start=SKIP_ROWS+1):
        if is_empty_row(row):
            continue

        r = dict(zip(cols, [clean(v) for v in row[:len(cols)]]))

        if not r["valeur"]:
            errors.append(f"  [briques] ligne {row_idx} : 'valeur' vide — ligne ignorée")
            continue

        brique = {
            "id":        r["id"] or make_id(),
            "categorie": r["categorie"] or "style",
            "valeur":    r["valeur"],
            "tags":      parse_tags(r["tags"]),
        }
        briques.append(brique)

    return briques, errors


def parse_csv(path):
    """Fallback CSV : colonnes = headers ligne 1, données ligne 2+."""
    import csv
    entries = []
    errors  = []
    cols_e  = ["id","titre","contenu","type","tags","notes","thumb_path","parent_id"]

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row_idx, row in enumerate(reader, start=2):
            titre = row.get("titre","").strip()
            if not titre:
                continue
            type_ = row.get("type","").strip()
            if type_ and type_ not in VALID_TYPES:
                errors.append(f"  [csv] ligne {row_idx} : type invalide '{type_}'")
                continue
            entries.append({
                "id":            row.get("id","").strip() or make_id(),
                "titre":         titre,
                "contenu":       row.get("contenu","").strip(),
                "type":          type_ or "prompt_texte",
                "tags":          parse_tags(row.get("tags","")),
                "notes":         row.get("notes","").strip(),
                "thumb_path":    row.get("thumb_path","").strip(),
                "date_creation": now_iso(),
                "parent_id":     row.get("parent_id","").strip() or None,
            })
    return entries, errors

# ── Main ─────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Import XLSX/CSV → import_queue.json")
    parser.add_argument("file", help="Fichier .xlsx ou .csv à importer")
    parser.add_argument("--replace", action="store_true",
                        help="action=replace (écrase tout) au lieu de merge (upsert)")
    args = parser.parse_args()

    src = Path(args.file)
    if not src.exists():
        sys.exit(f"Fichier introuvable : {src}")

    entries, briques, all_errors = [], [], []

    if src.suffix.lower() == ".csv":
        entries, errs = parse_csv(src)
        all_errors += errs
    else:
        wb = openpyxl.load_workbook(src, data_only=True)
        if "entries" in wb.sheetnames:
            e, errs = parse_entries_sheet(wb["entries"])
            entries += e; all_errors += errs
        if "briques" in wb.sheetnames:
            b, errs = parse_briques_sheet(wb["briques"])
            briques += b; all_errors += errs

    if all_errors:
        print(f"⚠️  {len(all_errors)} erreur(s) de validation :")
        for e in all_errors:
            print(e)

    queue = {
        "version":     "1.0",
        "exported_at": now_iso(),
        "action":      "replace" if args.replace else "merge",
        "entries":     entries,
        "briques":     briques,
    }

    OUTPUT_PATH.write_text(json.dumps(queue, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"✅  {len(entries)} entrée(s), {len(briques)} brique(s) → {OUTPUT_PATH}")
    if args.replace:
        print("    Mode REPLACE : les données existantes seront écrasées au prochain chargement.")
    else:
        print("    Mode MERGE : upsert par id (les entrées existantes sont mises à jour).")
    print("    → Recharge l'app dans le navigateur pour déclencher l'import.")

if __name__ == "__main__":
    main()
